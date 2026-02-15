from tally_pandl_export import export_pandl_from_tally
from ledger_sync import sync_ledgers_from_tally
from merge_header_footer import copy_all_parts
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
import pandas as pd
from pathlib import Path


# ==========================================================
# --- CONFIGURATION ---
# ==========================================================
base_dir = Path(__file__).parent.parent
xml_file = base_dir / "exports" / "PandL.xml"
template_file = base_dir / "config" / "template_kannada.xlsx"
mapping_file = base_dir / "config" / "ledger_mapping.xlsx"
output_file = base_dir / "output" / "body_PnL.xlsx"


# ==========================================================
# 1️⃣ Parse the Tally P&L XML section-wise
# ==========================================================
def parse_tally_xml():
    tree = ET.parse(xml_file)
    root = tree.getroot()

    income, expense = [], []
    current_section = None
    last_ledger = None

    for elem in root.iter():
        tag = elem.tag.upper()

        # Identify section headers
        if tag == "DSPDISPNAME":
            name_text = elem.text.strip() if elem.text else ""
            if name_text in ["Direct Incomes", "Indirect Incomes"]:
                current_section = "income"
                last_ledger = None
            elif name_text in ["Direct Expenses", "Indirect Expenses"]:
                current_section = "expense"
                last_ledger = None
            else:
                last_ledger = name_text  # ledger candidate

        elif tag == "BSSUBAMT" and last_ledger and current_section:
            amt_text = elem.text.strip() if elem.text else ""
            try:
                amt = float(amt_text)
            except:
                amt = 0.0
            if amt != 0:
                if current_section == "income":
                    income.append((last_ledger, abs(amt)))
                elif current_section == "expense":
                    expense.append((last_ledger, abs(amt)))
            last_ledger = None

    return income, expense


# ==========================================================
# 2️⃣ Translation and filtering
# ==========================================================
def load_mapping():
    mapping_df = pd.read_excel(mapping_file)
    mapping_df.columns = mapping_df.columns.str.strip()
    mapping_dict = {
        str(row["EnglishLedger"]).strip().lower(): str(row["KannadaLedger"]).strip()
        for _, row in mapping_df.iterrows()
        if pd.notna(row["EnglishLedger"]) and pd.notna(row["KannadaLedger"])
    }
    return mapping_dict


def translate_and_filter(data, mapping_dict):
    translated = []
    for name, amt in data:
        key = name.strip().lower()
        if key in mapping_dict and amt != 0:
            translated.append((mapping_dict[key], amt))
    return translated


# ==========================================================
# 3️⃣ Month-year in Kannada
# ==========================================================
def get_month_year_kn(date=None):
    """
    Return month and year in Kannada (e.g. "ಮಾರ್ಚ್ 2024").
    date: datetime or date instance; if None, uses datetime.now().
    """
    MONTHS_KN = {
        1: "ಜನವರಿ", 2: "ಫೆಬ್ರವರಿ", 3: "ಮಾರ್ಚ್", 4: "ಏಪ್ರಿಲ್", 5: "ಮೇ", 6: "ಜೂನ್",
        7: "ಜುಲೈ", 8: "ಆಗಸ್ಟ್", 9: "ಸೆಪ್ಟೆಂಬರ್", 10: "ಅಕ್ಟೋಬರ್", 11: "ನವೆಂಬರ್", 12: "ಡಿಸೆಂಬರ್"
    }
    dt = date if date is not None else datetime.now()
    return f"{MONTHS_KN[dt.month]} {dt.year}"


# ==========================================================
# 4️⃣ Excel utilities
# ==========================================================
def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _alignment_with_wrap(ref_cell):
    """Alignment from ref_cell with wrap_text=True."""
    a = ref_cell.alignment if ref_cell.has_style else None
    if a:
        return Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            text_rotation=a.text_rotation,
            wrap_text=True,
            shrink_to_fit=a.shrink_to_fit,
            indent=a.indent,
        )
    return Alignment(wrap_text=True)


def insert_data(ws, data, start_row, col_name, col_amt, ref_name, ref_amt):
    """
    Dynamically insert rows from start_row based on data length.
    Name column: 15-char width, text wrapped. Amount column: 7-char width.
    """
    if not data:
        return

    for i, (name_kn, amt) in enumerate(data):
        row = start_row + i
        c1 = ws.cell(row, col_name)
        c1.value = name_kn
        c2 = ws.cell(row, col_amt)
        c2.value = amt
        copy_style(ref_name, c1)
        copy_style(ref_amt, c2)
        c1.alignment = _alignment_with_wrap(ref_name)
        c2.number_format = u'₹ #,##0.00'


# ==========================================================
# 5️⃣ Generate Kannada P&L Excel body
# ==========================================================
def generate_kannada_pnl(income, expense, month_year_kn):
    wb = load_workbook(template_file)
    ws = wb.active

    # Replace $$monthYear$$ placeholder
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip() == "$$monthYear$$":
                cell.value = month_year_kn

    start_row = 2
    exp_name_col, exp_amt_col = 2, 3   # B, C
    inc_name_col, inc_amt_col = 5, 6   # E, F

    # Fixed column widths: name columns 15 chars (with wrap), amount columns 7 chars
    ws.column_dimensions[get_column_letter(exp_name_col)].width = 15
    ws.column_dimensions[get_column_letter(exp_amt_col)].width = 7
    ws.column_dimensions[get_column_letter(inc_name_col)].width = 15
    ws.column_dimensions[get_column_letter(inc_amt_col)].width = 7

    # reference styles
    ref_exp_name = ws.cell(start_row, exp_name_col)
    ref_exp_amt = ws.cell(start_row, exp_amt_col)
    ref_inc_name = ws.cell(start_row, inc_name_col)
    ref_inc_amt = ws.cell(start_row, inc_amt_col)

    # Insert expense & income
    insert_data(ws, expense, start_row, exp_name_col, exp_amt_col, ref_exp_name, ref_exp_amt)
    insert_data(ws, income, start_row, inc_name_col, inc_amt_col, ref_inc_name, ref_inc_amt)

    wb.save(output_file)
    print(f"✅ Kannada Profit & Loss generated successfully → {output_file}")
    print(f"   Income Ledgers: {len(income)} | Expense Ledgers: {len(expense)}")


# ==========================================================
# 6️⃣ Main execution flow
# ==========================================================
def main():
    report_date = None  # month/year for report; set from export "To" date
    print("Step 1: Export Profit & Loss XML from Tally")
    export_result = export_pandl_from_tally()
    if export_result is None:
        print("Skipping next steps (no XML exported).")
        return
    _, report_date = export_result

    print("\nStep 2: Sync ledgers before generating report")
    mapping_df = sync_ledgers_from_tally()
    if mapping_df is None:
        print("Skipping P&L generation (Tally not reachable).")
        return

    print("\nStep 3: Parse Profit & Loss XML")
    income, expense = parse_tally_xml()
    mapping_dict = load_mapping()
    income = translate_and_filter(income, mapping_dict)
    expense = translate_and_filter(expense, mapping_dict)

    print("\nStep 4: Generate Kannada Profit & Loss Excel body")
    month_year_kn = get_month_year_kn(report_date)
    generate_kannada_pnl(income, expense, month_year_kn)

    header_file = base_dir / "config" / "header_template.xlsx"
    body_file = base_dir / "output" / "body_PnL.xlsx"
    footer_file = base_dir / "config" / "footer_template.xlsx"
    final_file = base_dir / "output" / "final_PnL.xlsx"
    copy_all_parts(header_file, body_file, footer_file, final_file, month_year_kn=month_year_kn)

    print("\n All steps completed successfully!")


# ==========================================================
# Run script
# ==========================================================
if __name__ == "__main__":
    main()

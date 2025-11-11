# merge_header_footer_safe.py
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
from datetime import datetime

# --- File paths ---
base_dir = Path(__file__).parent.parent
header_file = base_dir / "config" /"header_template.xlsx"
body_file   = base_dir / "output" /"body_PnL.xlsx"
footer_file = base_dir / "config" /"footer_template.xlsx"
final_file  = base_dir / "output" /"final_PnL.xlsx"
header_with_month = base_dir / "output" /"header_with_month.xlsx"

def copy_sheet_to(ws_src, ws_dest, dest_start_row):
    """
    Copy all cells, styles, column widths and merged ranges from ws_src into ws_dest
    starting at row dest_start_row (1-based).
    Returns the number of rows copied.
    """
    max_row = ws_src.max_row
    max_col = ws_src.max_column

    # 1) Copy column widths for used columns (only once per dest column)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        src_dim = ws_src.column_dimensions.get(col_letter)
        if src_dim and src_dim.width is not None:
            # copy width directly to dest column (same letter)
            dest_letter = get_column_letter(col_idx)
            ws_dest.column_dimensions[dest_letter].width = src_dim.width

    # 2) Copy cells and styles
    for r in range(1, max_row + 1):
        dest_r = dest_start_row + r - 1
        for c in range(1, max_col + 1):
            src_cell = ws_src.cell(row=r, column=c)
            dest_cell = ws_dest.cell(row=dest_r, column=c, value=src_cell.value)

            # copy style if present
            if src_cell.has_style:
                try:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)
                except Exception:
                    # be resilient if some style attribute copying fails
                    pass

    # 3) Copy merged cell ranges (adjust row offsets)
    # ws_src.merged_cells.ranges is iterable of MergeCellRange objects
    for mr in ws_src.merged_cells.ranges:
        # mr.coord is like 'A1:C1' or 'B5:B7'
        min_col, min_row, max_col, max_row = mr.bounds  # bounds: (min_col, min_row, max_col, max_row)
        # compute new coordinates in dest sheet
        new_min_row = dest_start_row + (min_row - 1)
        new_max_row = dest_start_row + (max_row - 1)
        new_min_col = min_col
        new_max_col = max_col

        new_range = (
            f"{get_column_letter(new_min_col)}{new_min_row}:"
            f"{get_column_letter(new_max_col)}{new_max_row}"
        )

        try:
            ws_dest.merge_cells(new_range)
        except Exception:
            # ignore if merge fails (e.g., overlapping merges) but continue
            pass

    return max_row  # number of rows copied

def copy_all_parts(header_path, body_path, footer_path, output_path):
    wb_header = load_workbook(header_path)
    ws_header = wb_header.active
    # ==========================================================
    # 3️⃣ Month-year in Kannada
    # ==========================================================
    MONTHS_KN = {
        1:"ಜನವರಿ",2:"ಫೆಬ್ರವರಿ",3:"ಮಾರ್ಚ್",4:"ಏಪ್ರಿಲ್",5:"ಮೇ",6:"ಜೂನ್",
        7:"ಜುಲೈ",8:"ಆಗಸ್ಟ್",9:"ಸೆಪ್ಟೆಂಬರ್",10:"ಅಕ್ಟೋಬರ್",11:"ನವೆಂಬರ್",12:"ಡಿಸೆಂಬರ್"
    }
    now = datetime.now()
    month_year_kn = f"{MONTHS_KN[now.month]} {now.year}"
    for row in ws_header.iter_rows():
        for cell in row:
            if cell.value and "$$monthYear$$" in str(cell.value):
                new_text = str(cell.value).replace("$$monthYear$$", month_year_kn)
                cell.value = new_text
    wb_header.save(header_with_month)
    
    wb_header = load_workbook(header_with_month)
    ws_header = wb_header.active
    
    wb_body = load_workbook(body_path)
    ws_body = wb_body.active

    wb_footer = load_workbook(footer_path)
    ws_footer = wb_footer.active

    wb_final = Workbook()
    ws_final = wb_final.active

    current_row = 1

    # copy header
    rows_copied = copy_sheet_to(ws_header, ws_final, current_row)
    current_row += rows_copied

    # copy body
    rows_copied = copy_sheet_to(ws_body, ws_final, current_row)
    current_row += rows_copied

    # copy footer
    rows_copied = copy_sheet_to(ws_footer, ws_final, current_row)
    current_row += rows_copied

    # save final workbook
    wb_final.save(output_path)
    print(f"✅ Final merged file written to: {output_path}")

if __name__ == "__main__":
    copy_all_parts(header_file, body_file, footer_file, final_file)
